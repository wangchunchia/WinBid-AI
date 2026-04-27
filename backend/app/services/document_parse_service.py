import re
import subprocess
import zipfile
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from typing import Any
from uuid import uuid4

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.domain import DocumentChunk, SourceDocument


@dataclass
class ParsedChunkData:
    document_id: str
    page_no: int
    chunk_index: int
    chunk_type: str
    text_content: str
    chunk_id: str | None = None


@dataclass
class ParsedDocumentData:
    document_id: str
    file_name: str
    page_count: int
    chunks: list[ParsedChunkData]
    warnings: list[str]
    used_ocr: bool


class DocumentParseService:
    def __init__(self) -> None:
        self._repo_root = Path(__file__).resolve().parents[3]
        self._sentence_splitter = re.compile(r"(?<=[。；！？;?!])")
        self._leading_number_pattern = re.compile(r"^\s*[0-9一二三四五六七八九十百千]+[、.．\)]\s*")

    def parse_and_store_documents(
        self,
        db: Session,
        project_id: str,
        documents: list[SourceDocument],
    ) -> tuple[list[ParsedDocumentData], list[str]]:
        parsed_documents: list[ParsedDocumentData] = []
        all_warnings: list[str] = []

        for document in documents:
            try:
                parsed = self.parse_document(document)
            except Exception as exc:
                document.parse_status = "failed"
                all_warnings.append(f"{document.file_name}: parse failed: {exc}")
                continue

            self._replace_document_chunks(db, project_id, document.id, parsed.chunks)
            document.page_count = parsed.page_count
            document.parse_status = "parsed"
            parsed_documents.append(parsed)
            all_warnings.extend(parsed.warnings)
            if parsed.used_ocr:
                all_warnings.append(f"{document.file_name}: OCR fallback used on at least one page")

        return parsed_documents, all_warnings

    def list_document_chunks(self, db: Session, project_id: str, document_id: str) -> list[DocumentChunk]:
        return db.scalars(
            select(DocumentChunk)
            .where(DocumentChunk.project_id == project_id, DocumentChunk.document_id == document_id)
            .order_by(DocumentChunk.page_no.asc(), DocumentChunk.chunk_index.asc())
        ).all()

    def parse_document(self, document: SourceDocument) -> ParsedDocumentData:
        path = self._resolve_path(document.storage_uri)
        suffix = path.suffix.lower()

        if suffix == ".pdf":
            pages, warnings, used_ocr = self._parse_pdf(path)
        elif suffix in {".doc", ".docx"}:
            pages, warnings, used_ocr = self._parse_word(path)
        elif suffix in {".xlsx", ".xls"}:
            pages, warnings, used_ocr = self._parse_spreadsheet(path)
        elif suffix in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}:
            pages, warnings, used_ocr = self._parse_image(path)
        elif suffix in {".txt", ".md"}:
            pages, warnings, used_ocr = [self._normalize_text(path.read_text(encoding="utf-8", errors="ignore"))], [], False
        else:
            raise ValueError(f"Unsupported document type: {suffix}")

        chunks = self._build_chunks(document.id, pages)
        if not chunks:
            warnings.append(f"{document.file_name}: no extractable text found")

        return ParsedDocumentData(
            document_id=document.id,
            file_name=document.file_name,
            page_count=max(len(pages), 1),
            chunks=chunks,
            warnings=warnings,
            used_ocr=used_ocr,
        )

    def _replace_document_chunks(
        self,
        db: Session,
        project_id: str,
        document_id: str,
        chunks: list[ParsedChunkData],
    ) -> None:
        db.execute(delete(DocumentChunk).where(DocumentChunk.document_id == document_id))
        for chunk in chunks:
            chunk_id = str(uuid4())
            chunk.chunk_id = chunk_id
            db.add(
                DocumentChunk(
                    id=chunk_id,
                    project_id=project_id,
                    document_id=document_id,
                    page_no=chunk.page_no,
                    chunk_index=chunk.chunk_index,
                    chunk_type=chunk.chunk_type,
                    text_content=chunk.text_content,
                    char_count=len(chunk.text_content),
                )
            )

    def _resolve_path(self, storage_uri: str) -> Path:
        if storage_uri.startswith("file://"):
            path = Path(storage_uri.removeprefix("file://"))
        else:
            path = Path(storage_uri)

        if path.is_absolute():
            resolved = path
        else:
            resolved = (self._repo_root / path).resolve()

        if not resolved.exists():
            raise FileNotFoundError(f"Document not found: {storage_uri}")
        return resolved

    def _parse_pdf(self, path: Path) -> tuple[list[str], list[str], bool]:
        try:
            import fitz
        except ImportError as exc:
            raise RuntimeError("PyMuPDF is required for PDF parsing") from exc

        warnings: list[str] = []
        used_ocr = False
        pages: list[str] = []

        with fitz.open(path) as pdf:
            for page_index in range(pdf.page_count):
                page = pdf.load_page(page_index)
                text = self._normalize_text(page.get_text("text"))
                if not text:
                    ocr_text = self._ocr_pdf_page(page)
                    if ocr_text:
                        text = ocr_text
                        used_ocr = True
                    else:
                        warnings.append(f"{path.name}: page {page_index + 1} has no selectable text and OCR was unavailable")
                pages.append(text)

        return pages, warnings, used_ocr

    def _ocr_pdf_page(self, page: Any) -> str:
        ocr = self._get_ocr_runtime()
        if not ocr:
            return ""

        import fitz

        pytesseract, image_module = ocr
        pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        image = image_module.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)
        return self._normalize_text(pytesseract.image_to_string(image, lang="chi_sim+eng"))

    def _parse_word(self, path: Path) -> tuple[list[str], list[str], bool]:
        text = self._extract_with_textutil(path)
        warnings: list[str] = []
        if not text and path.suffix.lower() == ".docx":
            text = self._extract_docx_xml(path)
            warnings.append(f"{path.name}: fell back to DOCX XML parsing")
        if not text:
            raise ValueError("No extractable text found in Word document")
        return [text], warnings, False

    def _parse_spreadsheet(self, path: Path) -> tuple[list[str], list[str], bool]:
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            return self._parse_xlsx(path), [], False
        return self._parse_xls(path), [], False

    def _parse_image(self, path: Path) -> tuple[list[str], list[str], bool]:
        ocr = self._get_ocr_runtime()
        if not ocr:
            raise RuntimeError("pytesseract and local tesseract are required for image OCR")
        pytesseract, image_module = ocr
        image = image_module.open(path)
        text = self._normalize_text(pytesseract.image_to_string(image, lang="chi_sim+eng"))
        return [text], [], True

    def _extract_with_textutil(self, path: Path) -> str:
        try:
            completed = subprocess.run(
                ["textutil", "-convert", "txt", "-stdout", str(path)],
                check=False,
                capture_output=True,
                text=True,
            )
        except FileNotFoundError:
            return ""

        if completed.returncode != 0:
            return ""
        return self._normalize_text(completed.stdout)

    def _extract_docx_xml(self, path: Path) -> str:
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml").decode("utf-8", errors="ignore")
        text = re.sub(r"</w:p>", "\n", xml)
        text = re.sub(r"<[^>]+>", "", text)
        return self._normalize_text(unescape(text))

    def _parse_xlsx(self, path: Path) -> list[str]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for .xlsx parsing") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        pages: list[str] = []
        for sheet in workbook.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell not in (None, "")]
                if values:
                    rows.append(" | ".join(values))
            pages.append(self._normalize_text(f"{sheet.title}\n" + "\n".join(rows)))
        return pages or [""]

    def _parse_xls(self, path: Path) -> list[str]:
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("xlrd is required for .xls parsing") from exc

        workbook = xlrd.open_workbook(path)
        pages: list[str] = []
        for sheet in workbook.sheets():
            rows: list[str] = []
            for row_idx in range(sheet.nrows):
                values = [str(value).strip() for value in sheet.row_values(row_idx) if str(value).strip()]
                if values:
                    rows.append(" | ".join(values))
            pages.append(self._normalize_text(f"{sheet.name}\n" + "\n".join(rows)))
        return pages or [""]

    def _get_ocr_runtime(self) -> tuple[Any, Any] | None:
        try:
            import pytesseract
            from PIL import Image
        except ImportError:
            return None

        try:
            completed = subprocess.run(
                ["tesseract", "--version"],
                check=False,
                capture_output=True,
                text=True,
            )
        except FileNotFoundError:
            return None

        if completed.returncode != 0:
            return None
        return pytesseract, Image

    def _build_chunks(self, document_id: str, pages: list[str]) -> list[ParsedChunkData]:
        chunks: list[ParsedChunkData] = []
        seen_exact_blocks: set[str] = set()
        seen_fingerprints: set[str] = set()
        for page_no, page_text in enumerate(pages, start=1):
            if not page_text.strip():
                continue
            paragraphs = [part.strip() for part in re.split(r"\n{2,}", page_text) if part.strip()]
            chunk_index = 1
            for paragraph in paragraphs:
                for block in self._split_long_text(paragraph):
                    normalized_block = self._normalize_text(block)
                    if not normalized_block:
                        continue
                    fingerprint = self._build_dedupe_fingerprint(normalized_block)
                    if normalized_block in seen_exact_blocks:
                        continue
                    if fingerprint and fingerprint in seen_fingerprints and len(normalized_block) >= 24:
                        continue
                    seen_exact_blocks.add(normalized_block)
                    if fingerprint:
                        seen_fingerprints.add(fingerprint)
                    chunks.append(
                        ParsedChunkData(
                            document_id=document_id,
                            page_no=page_no,
                            chunk_index=chunk_index,
                            chunk_type="paragraph",
                            text_content=normalized_block,
                        )
                    )
                    chunk_index += 1
        return chunks

    def _split_long_text(self, text: str, max_chars: int = 800) -> list[str]:
        normalized = self._normalize_text(text)
        if len(normalized) <= max_chars:
            return [normalized]

        parts = [part.strip() for part in self._sentence_splitter.split(normalized) if part.strip()]
        blocks: list[str] = []
        current = ""

        for part in parts:
            candidate = f"{current}{part}"
            if current and len(candidate) > max_chars:
                blocks.append(current.strip())
                current = part
            else:
                current = candidate

        if current.strip():
            blocks.append(current.strip())

        return blocks or [normalized]

    def _normalize_text(self, text: str) -> str:
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text.strip()

    def _build_dedupe_fingerprint(self, text: str) -> str:
        normalized = self._leading_number_pattern.sub("", text)
        normalized = re.sub(r"\s+", "", normalized)
        normalized = re.sub(r"[0-9０-９]", "", normalized)
        normalized = re.sub(r"[，,。；;：:、（）()【】\\[\\]《》“”\"'‘’·/\\\\_-]", "", normalized)
        return normalized[:160]


document_parse_service = DocumentParseService()
